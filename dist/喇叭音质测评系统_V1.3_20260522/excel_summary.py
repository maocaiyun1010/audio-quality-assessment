# -*- coding: utf-8 -*-
"""
音效评测汇总 Excel：生成 .xlsx，用 Excel 打开即为表格（非纯文本 TSV）。
依赖：openpyxl
"""
from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from markdown_report import DIMENSION_KEYS, _dim_numeric

from tsv_report import (
    SCALE_PAIRWISE,
    SCALE_SINGLE,
    _row_mean_five,
    comparison_overall_narrative,
    single_overall_narrative,
)


def write_evaluation_xlsx(
    path: str | Path,
    *,
    test_name: str,
    test_device: str,
    ref_device: str,
    eval_model_name: str = "",
    comparison_mode: bool,
    cross_session: bool,
    rows: list[dict[str, Any]],
    dim_avgs: dict[str, float | None],
    grand: float | None,
    dut_label: str = "",
    ref_label: str = "",
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("请安装 openpyxl：pip install openpyxl") from exc

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "音效评测汇总"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    meta_label_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center_num = Alignment(horizontal="center", vertical="center")

    r = 1
    mode = (
        "跨会话刺激比较（五维整数−3～+3，正为被测优）"
        if cross_session
        else (
            "同会话刺激比较（五维整数−3～+3，正为被测优）"
            if comparison_mode
            else "单终端绝对分（五维1～10）"
        )
    )
    meta_rows: list[tuple[str, str]] = [
        ("报告标题", test_name),
    ]
    _em = (eval_model_name or "").strip()
    if _em:
        meta_rows.append(("评测大模型", _em))
    meta_rows.extend(
        [
            ("被测/主测标识", test_device),
            ("对比/参考信息", ref_device),
            ("评测模式", mode),
        ]
    )
    for label, val in meta_rows:
        ws.cell(r, 1, label).font = meta_label_font
        ws.cell(r, 2, val).alignment = wrap
        r += 1

    r += 1  # 空行

    hdr = ["测评音频"] + list(DIMENSION_KEYS) + ["本行五维均分"]
    if not comparison_mode:
        hdr.append("综合分")
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(r, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    data_start = r

    for row in rows:
        ws.cell(r, 1, str(row.get("节目") or "")).alignment = wrap
        for ci, k in enumerate(DIMENSION_KEYS, start=2):
            v = _dim_numeric(row, k)
            if v is None:
                ws.cell(r, ci, "—").alignment = center_num
            else:
                if comparison_mode and abs(v - round(v)) < 1e-9:
                    ws.cell(r, ci, int(round(v))).alignment = center_num
                else:
                    ws.cell(r, ci, round(v, 2)).alignment = center_num
        rm = _row_mean_five(row)
        c_rm = 2 + len(DIMENSION_KEYS)
        ws.cell(r, c_rm, round(rm, 2) if rm is not None else "—").alignment = center_num
        if not comparison_mode:
            z = _dim_numeric(row, "综合分")
            if z is None:
                ws.cell(r, c_rm + 1, "—").alignment = center_num
            else:
                ws.cell(r, c_rm + 1, round(z, 2)).alignment = center_num
        r += 1

    # 汇总行
    foot_font = Font(bold=True)
    foot_fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(r, 1, "各维度列平均及总平均").font = foot_font
    ws.cell(r, 1).fill = foot_fill
    for ci, k in enumerate(DIMENSION_KEYS, start=2):
        a = dim_avgs.get(k)
        cell = ws.cell(r, ci, round(a, 2) if a is not None else "—")
        cell.font = foot_font
        cell.fill = foot_fill
        cell.alignment = center_num
    c_rm = 2 + len(DIMENSION_KEYS)
    if grand is None or (isinstance(grand, float) and math.isnan(grand)):
        grand_cell_val: Any = "—"
    else:
        grand_cell_val = round(float(grand), 2)
    cell_grand = ws.cell(r, c_rm, grand_cell_val)
    cell_grand.font = foot_font
    cell_grand.fill = foot_fill
    cell_grand.alignment = center_num
    if not comparison_mode:
        znums = [_dim_numeric(row, "综合分") for row in rows]
        znums = [x for x in znums if x is not None]
        zavg = round(sum(znums) / len(znums), 2) if znums else None
        c_z = c_rm + 1
        cell_z = ws.cell(r, c_z, round(zavg, 2) if zavg is not None else "—")
        cell_z.font = foot_font
        cell_z.fill = foot_fill
        cell_z.alignment = center_num
    r += 1

    r += 1
    scale = SCALE_PAIRWISE if comparison_mode else SCALE_SINGLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hdr))
    c = ws.cell(r, 1, scale)
    c.alignment = wrap
    c.font = Font(size=10, color="333333")
    r += 2

    dl = (dut_label or test_device or "被测").strip()
    rl = (ref_label or "对比样机").strip()
    if comparison_mode:
        narr = comparison_overall_narrative(dim_avgs, dut_label=dl, ref_label=rl)
    else:
        narr = single_overall_narrative(grand, len(rows))

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hdr))
    cell_n = ws.cell(r, 1, narr)
    cell_n.alignment = wrap
    cell_n.font = Font(size=10)
    ws.row_dimensions[r].height = 72

    # 列宽
    ws.column_dimensions["A"].width = 52
    for i in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = f"A{data_start}"
    wb.save(str(p))
