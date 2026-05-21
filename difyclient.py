import json
import os
import re
import shutil
import sys
import tempfile
import time
import uuid
import mimetypes
from typing import Any, Callable, Optional, Sequence
from pathlib import Path

import requests

# 音频上传与评分流式请求：首次失败后可重试（共 4 次尝试 = 1 次初始 + 重试 3 次）
DIFY_RETRY_MAX_ATTEMPTS = 4
DIFY_RETRY_INTERVAL_SEC = 5.0

_DEFAULT_UPLOAD_CONNECT = 25.0
_DEFAULT_UPLOAD_READ = 240.0
_DEFAULT_CHAT_CONNECT = 30.0
_DEFAULT_CHAT_READ = 120.0
_DEFAULT_STREAM_WALL = 480.0
_DEFAULT_STREAM_PROGRESS_SEC = 30.0
_DEFAULT_DIFY_UPLOAD_MAX_AUDIO_SEC = 60.0


def dify_upload_max_audio_seconds() -> float | None:
    """
    上传 Dify 前对音频的最长保留秒数（仅取开头一段；本地录音文件不改动）。

    - 环境变量 ``DIFY_UPLOAD_MAX_AUDIO_SECONDS``：正数 = 上限秒数；``0`` = 不截断、整文件上传。
    - 未设置时默认 **60** 秒。
    """
    raw = (os.environ.get("DIFY_UPLOAD_MAX_AUDIO_SECONDS") or "").strip()
    if not raw:
        return _DEFAULT_DIFY_UPLOAD_MAX_AUDIO_SEC
    try:
        v = float(raw)
    except ValueError:
        return _DEFAULT_DIFY_UPLOAD_MAX_AUDIO_SEC
    if v <= 0.0:
        return None
    return v


def _float_env(name: str, default: float) -> float:
    """从环境变量读取浮点秒数；非法或空则返回 default。"""
    raw = os.environ.get(name, "").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def _env_truthy(name: str) -> bool:
    return (os.environ.get(name) or "").strip().lower() in ("1", "true", "yes", "on")


def _agent_debug_enabled() -> bool:
    """遗留 agent 调试日志默认关闭，避免导入或评分时无意写盘。"""
    return _env_truthy("SPEAKER_AGENT_DEBUG_LOG")


def _int_env(name: str, default: int) -> int:
    raw = os.environ.get(name, "").strip()
    if not raw:
        return default
    try:
        return int(raw)
    except ValueError:
        return default


# Dify 侧常见限制：``audio_eval_prompt`` 须 **小于 256 个字符**（本仓库默认按最多 255 个码点截断）。
# 完整评分指令、JSON Schema 仍在 HTTP 请求体的 ``query`` 正文中，不在此字段重复长文。
_DIFY_AUDIO_EVAL_PROMPT_MAX_LEN = 255
_DEFAULT_AUDIO_EVAL_PROMPT_SHORT = (
    "请依据本消息正文中的评测规则与 JSON Schema，结合附件音频完成评测；"
    "仅输出要求的一个 JSON 对象，勿 Markdown 围栏，勿拒答。"
)


def _audio_eval_prompt_max_len() -> int:
    n = _int_env("DIFY_AUDIO_EVAL_PROMPT_MAX_LEN", _DIFY_AUDIO_EVAL_PROMPT_MAX_LEN)
    return max(1, min(n, 2000))


def _clip_dify_audio_eval_prompt(s: str) -> str:
    mx = _audio_eval_prompt_max_len()
    t = (s or "").strip()
    if not t:
        return ""
    if len(t) <= mx:
        return t
    if mx <= 1:
        return t[:mx]
    return t[: mx - 1] + "…"


# 侧栏展示名 → 火山 Ark / 部分 Dify 节点实际识别的 model 字符串（大小写与连字符常不一致）。
# 若贵司控制台 Endpoint 使用其它 ID（如带日期的 ep- 或 doubao-seed-2-0-xxxx），请在项目根
# ``web_ui_provider_model_map.json`` 覆盖（见 ``_merged_provider_model_aliases``）。
_BUILTIN_PROVIDER_MODEL_ALIASES: dict[str, str] = {
    "Doubao-Seed-2.0-pro": "doubao-seed-2.0-pro",
    "Doubao-Seed-2.0-Lite": "doubao-seed-2.0-lite",
}

_provider_alias_map_cache: tuple[float | None, dict[str, str]] | None = None


def _merged_provider_model_aliases() -> dict[str, str]:
    """合并内置别名与可选 ``web_ui_provider_model_map.json``（用户键覆盖同键内置值）。"""
    global _provider_alias_map_cache
    custom_path = (os.environ.get("SPEAKER_PROVIDER_MODEL_MAP_PATH") or "").strip()
    path = Path(custom_path) if custom_path else Path(__file__).resolve().parent / "web_ui_provider_model_map.json"
    mtime: float | None = None
    try:
        mtime = path.stat().st_mtime if path.is_file() else None
    except OSError:
        mtime = None
    if _provider_alias_map_cache is not None and _provider_alias_map_cache[0] == mtime:
        return _provider_alias_map_cache[1]
    merged: dict[str, str] = dict(_BUILTIN_PROVIDER_MODEL_ALIASES)
    if path.is_file():
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                for k, v in raw.items():
                    ks = str(k).strip()
                    vs = str(v).strip()
                    if ks.startswith("_") or not ks or not vs:
                        continue
                    merged[ks] = vs
        except Exception:
            pass
    _provider_alias_map_cache = (mtime, merged)
    return merged


def resolve_selected_model_for_dify_inputs(model_name: str) -> str:
    """
    将侧栏 / 环境变量中的展示名转为 Dify ``inputs.selected_model`` 建议传入的提供商 model 串。

    设 ``SPEAKER_DISABLE_PROVIDER_MODEL_ALIAS=1`` 时原样返回（兼容已在 Dify 内写死展示名且与上游一致的场景）。
    """
    m = (model_name or "").strip()
    if not m or _env_truthy("SPEAKER_DISABLE_PROVIDER_MODEL_ALIAS"):
        return m
    aliases = _merged_provider_model_aliases()
    return aliases.get(m, m)


def _dify_inputs_audio_eval_prompt(_query: str, *, explicit: Optional[str] = None) -> dict[str, str]:
    """
    若 Dify Chat 应用在「开始表单」声明必填变量 ``audio_eval_prompt``，须在 ``inputs`` 中传入。

    该字段在常见部署下 **长度上限 <256**（程序默认截断至 ``DIFY_AUDIO_EVAL_PROMPT_MAX_LEN``，缺省 255），
    故**不能**把与 ``query`` 相同的长提示词整段塞入。详细规则请放在 ``query`` 中。

    取值优先级：``explicit``（web_ui_prompt_overrides.json）> 环境变量 ``DIFY_AUDIO_EVAL_PROMPT`` >
    内置短文 ``_DEFAULT_AUDIO_EVAL_PROMPT_SHORT``。

    设 ``DIFY_OMIT_AUDIO_EVAL_PROMPT_INPUT=1`` 可关闭传入（兼容未声明该变量的应用）。
    """
    if _env_truthy("DIFY_OMIT_AUDIO_EVAL_PROMPT_INPUT"):
        return {}
    v = (explicit or "").strip()
    if not v:
        v = (os.environ.get("DIFY_AUDIO_EVAL_PROMPT") or "").strip()
    if not v:
        v = _DEFAULT_AUDIO_EVAL_PROMPT_SHORT
    v = _clip_dify_audio_eval_prompt(v)
    return {"audio_eval_prompt": v} if v else {}


def _dify_inputs_selected_model(*, explicit: Optional[str] = None) -> dict[str, str]:
    """
    部分 Dify 应用在开始表单中要求 ``selected_model``。

    优先级：``explicit``（``web_ui_prompt_overrides.json``）> ``DIFY_SELECTED_MODEL`` >
    ``SPEAKER_EVAL_MODEL_NAME``（Web UI 侧栏选项会写入该环境变量）。

    写入前会经 ``resolve_selected_model_for_dify_inputs`` 做可选别名映射（解决火山侧
    ``Model Doubao-Seed-2.0-pro not exist`` 等大小写/命名不一致问题）。

    ``DIFY_OMIT_SELECTED_MODEL_INPUT=1``：不传该字段（兼容未声明该变量的应用）。
    """
    if _env_truthy("DIFY_OMIT_SELECTED_MODEL_INPUT"):
        return {}
    v = (explicit or "").strip()
    if not v:
        v = (os.environ.get("DIFY_SELECTED_MODEL") or "").strip()
    if not v:
        v = (os.environ.get("SPEAKER_EVAL_MODEL_NAME") or "").strip()
    v = resolve_selected_model_for_dify_inputs(v)
    return {"selected_model": v} if v else {}


def _dify_chat_form_inputs(
    query: str,
    *,
    audio_eval_prompt_explicit: Optional[str] = None,
    selected_model_explicit: Optional[str] = None,
    omit_audio_eval_prompt: bool = False,
) -> dict[str, str]:
    """合并 ``audio_eval_prompt`` 与 ``selected_model`` 等开始表单变量。"""
    out: dict[str, str] = {}
    if not omit_audio_eval_prompt:
        out.update(_dify_inputs_audio_eval_prompt(query, explicit=audio_eval_prompt_explicit))
    out.update(_dify_inputs_selected_model(explicit=selected_model_explicit))
    return out


def _file_upload_timeout() -> tuple[float, float]:
    """返回 (connect, read) 秒数，供 ``requests.post(..., files=...)`` 使用。"""
    c = _float_env("DIFY_FILE_UPLOAD_CONNECT_TIMEOUT_SEC", _DEFAULT_UPLOAD_CONNECT)
    r = _float_env("DIFY_FILE_UPLOAD_READ_TIMEOUT_SEC", _DEFAULT_UPLOAD_READ)
    return (max(5.0, c), max(30.0, r))


def _chat_request_timeout(fallback_total: float) -> float | tuple[float, float]:
    """
    对话/流式请求的 ``requests`` 超时 (connect, read)。

    read 为**相邻两次收到数据**之间的最长等待。默认 (30s, 120s)，避免旧版单值 600s
    导致「一直无结果、像卡死」。若需更长等待可设 ``DIFY_CHAT_READ_TIMEOUT_SEC=300``。
    设 ``DIFY_CHAT_USE_LEGACY_SINGLE_TIMEOUT=1`` 可恢复为单个 fallback_total 秒数。
    """
    if _env_truthy("DIFY_CHAT_USE_LEGACY_SINGLE_TIMEOUT"):
        return max(30.0, float(fallback_total))
    c = _float_env("DIFY_CHAT_CONNECT_TIMEOUT_SEC", _DEFAULT_CHAT_CONNECT)
    r = _float_env("DIFY_CHAT_READ_TIMEOUT_SEC", _DEFAULT_CHAT_READ)
    if not (os.environ.get("DIFY_CHAT_READ_TIMEOUT_SEC") or "").strip():
        r = min(float(fallback_total), _DEFAULT_CHAT_READ)
    return (max(5.0, c), max(15.0, r))


def _stream_collect_wall_sec() -> float:
    return max(60.0, _float_env("DIFY_STREAM_COLLECT_MAX_SEC", _DEFAULT_STREAM_WALL))


def _stream_progress_interval_sec() -> float:
    return max(10.0, _float_env("DIFY_STREAM_PROGRESS_INTERVAL_SEC", _DEFAULT_STREAM_PROGRESS_SEC))


def _empty_body_stream_attempts() -> int:
    return max(1, min(3, _int_env("DIFY_EMPTY_BODY_STREAM_ATTEMPTS", 2)))


def _blocking_max_attempts() -> int:
    return max(1, min(DIFY_RETRY_MAX_ATTEMPTS, _int_env("DIFY_BLOCKING_MAX_ATTEMPTS", 2)))


def _live_scoring_detail(detail: str) -> None:
    try:
        from live_eval_log import append_live_scoring_detail

        append_live_scoring_detail(detail)
    except Exception:
        pass


# 多附件上传后若立即发 chat，Dify 侧偶发未绑定附件，模型报「无法解析附件」；可通过环境变量微调。
# DIFY_PER_AUDIO_UPLOAD_GAP_SEC：每路文件上传成功后的间隔（秒）
# DIFY_STIMULUS_FIRST_TRACK_GAP_SEC：见 scoring.py，首条刺激比较评分前的附加等待（秒）
# DIFY_STIMULUS_EARLY_TRACK_RETRY_COOLDOWN_SEC：见 scoring.py，音源 1/2 整轨失败后的冷却再评（秒，0=关闭；旧名 TRACK1_RETRY 仍可读）
# DIFY_STIMULUS_POST_UPLOAD_CHAT_DELAY_SEC：全部上传完成后、发起评分对话前的等待（秒）
# DIFY_SINGLE_AUDIO_POST_UPLOAD_DELAY_SEC：单文件 analyze_audio 上传完成后等待（秒）
# DIFY_FILE_UPLOAD_CONNECT_TIMEOUT_SEC / DIFY_FILE_UPLOAD_READ_TIMEOUT_SEC：multipart 上传（秒），
#   避免网关半开时整段 300s 无返回导致界面长期停在「上传中」；大文件可把 read 调大。
# DIFY_CHAT_CONNECT_TIMEOUT_SEC / DIFY_CHAT_READ_TIMEOUT_SEC：评分对话（秒），read 为相邻数据块间隔上限。
# DIFY_STREAM_COLLECT_MAX_SEC：单次流式正文采集总墙钟上限（秒）。
# DIFY_STREAM_PROGRESS_INTERVAL_SEC：等待模型时向日志/实时步骤条刷新进度的间隔（秒）。
# DIFY_EMPTY_BODY_STREAM_ATTEMPTS：流式无正文时的重复次数（默认 2，可设为 1 以更快失败）。
# DIFY_BLOCKING_MAX_ATTEMPTS：blocking 兜底最多尝试次数（默认 2）。
# DIFY_SKIP_BLOCKING_FALLBACK=1：流式无正文时不走 blocking（更快失败，排查用）。
_DEFAULT_PER_UPLOAD_GAP = 0.65
_DEFAULT_STIMULUS_POST_UPLOAD = 5.0
_DEFAULT_SINGLE_POST_UPLOAD = 1.0

# #region agent log
_DEBUG_UPLOAD_LOG = Path(__file__).resolve().parent / "debug-0d224e.log"


def _agent_dbg_log_paths() -> list[Path]:
    """多路径写入：避免子进程 cwd 与仓库不一致时找不到日志。"""
    seen: set[str] = set()
    out: list[Path] = []
    _paths: list[Path] = [
        _DEBUG_UPLOAD_LOG,
        Path.cwd() / "debug-0d224e.log",
        Path(tempfile.gettempdir()) / "debug-0d224e.log",
    ]
    try:
        from config import ANALYSIS_DIR  # type: ignore

        _paths.append(Path(ANALYSIS_DIR) / "debug-0d224e.log")
    except Exception:
        pass
    for p in _paths:
        try:
            key = str(p.resolve())
        except Exception:
            key = str(p)
        if key not in seen:
            seen.add(key)
            out.append(p)
    return out


def _agent_dbg_upload(hypothesis_id: str, location: str, message: str, data: dict) -> None:
    if not _agent_debug_enabled():
        return
    rec = {
        "sessionId": "0d224e",
        "timestamp": int(time.time() * 1000),
        "hypothesisId": hypothesis_id,
        "location": location,
        "message": message,
        "data": data,
    }
    line = json.dumps(rec, ensure_ascii=False) + "\n"
    _ok_any = False
    _first_err: str | None = None
    for _lp in _agent_dbg_log_paths():
        try:
            _lp.parent.mkdir(parents=True, exist_ok=True)
            with open(_lp, "a", encoding="utf-8") as _f:
                _f.write(line)
                _f.flush()
            _ok_any = True
        except Exception as exc:
            if _first_err is None:
                _first_err = f"{type(exc).__name__}: {exc}"
    if not _ok_any and _first_err:
        try:
            sys.stderr.write(f"[debug-0d224e] 无法写入日志文件: {_first_err}\n")
        except Exception:
            pass


# #endregion


def first_balanced_json_object_slice(s: str, start_at: int = 0) -> Optional[str]:
    """
    从 ``start_at`` 起找到第一个 ``{``，切出与之平衡的顶层 ``{...}``（尊重字符串内的引号与转义）。

    模型偶发在同一回复里输出两个完全相同的 JSON 对象（中间无逗号），若用 ``find('{')`` + ``rfind('}')``
    会得到非法 JSON；评分与展示应只取第一段。

    另：SSE/工作流可能先输出用量元数据 ``{...}`` 再输出评分 JSON；需从后续 ``{`` 重新切片（见
    ``iter_balanced_json_object_slices``）。
    """
    if not s:
        return None
    i = s.find("{", max(0, start_at))
    if i < 0:
        return None
    depth = 0
    in_str = False
    esc = False
    for j in range(i, len(s)):
        ch = s[j]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return s[i : j + 1]
    return None


def iter_balanced_json_object_slices(
    s: str, *, start_at: int = 0, max_objects: int = 96
) -> list[str]:
    """
    枚举文本中从各 ``{`` 起可解析的顶层 ``{...}`` 子串（去重）。

    用于粘连 ``{usage...}{评分...}``、或正文前缀非 JSON 时仍能命中含「五维」键的对象。
    """
    if not s:
        return []
    out: list[str] = []
    seen: set[str] = set()
    pos = max(0, start_at)
    while len(out) < max_objects:
        i = s.find("{", pos)
        if i < 0:
            break
        chunk = first_balanced_json_object_slice(s, start_at=i)
        if chunk and chunk not in seen:
            seen.add(chunk)
            out.append(chunk)
        pos = i + 1
    return out


def _pretty_json_text_for_display(text: Optional[str]) -> str:
    """若正文为合法 JSON（可含 Markdown 代码围栏），缩进排版后输出，便于控制台阅读。"""
    if not text:
        return text or ""
    s = text.strip()
    if s.startswith("```"):
        lines = s.splitlines()
        if lines and lines[0].lstrip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    try:
        return json.dumps(json.loads(s), ensure_ascii=False, indent=2)
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    chunk = first_balanced_json_object_slice(s)
    if chunk:
        try:
            return json.dumps(json.loads(chunk), ensure_ascii=False, indent=2)
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    return text


def _stimulus_compare_runtime_block(
    *,
    stimulus_label: str,
    device_slot_labels: Sequence[str],
    comparison_variant: str,
    dut_attachment_index: int,
    ref_attachment_index: int,
) -> str:
    """仅含本轮事实（音源、附件顺序），不含评分规则长文；与「final」自定义提示词搭配。"""
    mapping_lines = "\n".join(
        f"- 第 {i + 1} 个音频附件：**{device_slot_labels[i]}**"
        for i in range(len(device_slot_labels))
    )
    dut_n = int(dut_attachment_index) + 1
    ref_n = int(ref_attachment_index) + 1
    variant = (comparison_variant or "same_session").strip().lower()
    cross_note = ""
    if variant == "cross_session":
        cross_note = (
            "（跨会话：两路附件来自两次独立测试会话，同一麦克风链路分时采集。）\n"
        )
    return (
        f"【本轮任务上下文 · 事实信息】\n"
        f"{cross_note}"
        f"- 当前音源标识：{stimulus_label}\n"
        f"- 附件与槽位（按顺序解读）：\n{mapping_lines}\n"
        f"- 约定：第 {dut_n} 个附件 = 被测设备，第 {ref_n} 个附件 = 对比设备。\n"
        "- 附件为已上传的完整 WAV，须基于音频听感评分；禁止以无法试听为由拒答或五维全 0 占位。"
    )


def _build_stimulus_compare_query(
    *,
    extra_instruction: str,
    stimulus_label: str,
    device_slot_labels: Sequence[str],
    comparison_variant: str,
    dut_attachment_index: int,
    ref_attachment_index: int,
    prompt_mode: str = "builtin",
) -> str:
    """
    构造刺激比较（双路）Dify 对话 ``query`` 全文。

    ``prompt_mode="final"``：仅发送自定义提示词 + 本轮上下文；``builtin``：拼接内置长模板（旧行为）。
    """
    runtime = _stimulus_compare_runtime_block(
        stimulus_label=stimulus_label,
        device_slot_labels=device_slot_labels,
        comparison_variant=comparison_variant,
        dut_attachment_index=dut_attachment_index,
        ref_attachment_index=ref_attachment_index,
    )
    body = (extra_instruction or "").strip()
    mode = (prompt_mode or "builtin").strip().lower()
    if mode == "final" and body:
        return f"{body}\n\n{runtime}"
    if mode == "final" and not body:
        mode = "builtin"

    mapping_lines = "\n".join(
        f"- 第 {i + 1} 个音频附件对应设备槽位：**{device_slot_labels[i]}**"
        for i in range(len(device_slot_labels))
    )
    stim_json = json.dumps(stimulus_label, ensure_ascii=False)
    di = int(dut_attachment_index)
    ri = int(ref_attachment_index)
    dut_n, ref_n = di + 1, ri + 1

    variant = (comparison_variant or "same_session").strip().lower()
    cross_block = ""
    intro_line = (
        "你是资深电声与音频测试工程师。以下附件为**同一标准测试音源**分别在【被测设备】与【对比设备】"
        "内置喇叭上外放后，经**同一麦克风采集链路**录制的文件。请对比两路录音所反映的**内置喇叭主观听感差异**。"
    )
    if variant == "cross_session":
        cross_block = """
【跨会话说明】两路附件来自**两次独立测试会话**：每次仅连接一台终端，在**同一支麦克风、同一条采集链路**下，于**尽量相同的摆位与房间声学条件**中**分时**录制（非两台同时外放混录）。
评分任务不变：仍是在**同一音源刺激**下，依据主观听感比较【被测】相对【对比】的内置喇叭表现。
**下列「评分原则」与「JSON Schema」与同一会话内顺序双机录制的情形完全一致**；不得因跨会话改用 1～10 分、不得省略任一维度分差、不得改变 -3～+3 的整数约束。

"""
        intro_line = (
            "你是资深电声与音频测试工程师。以下两路附件为**跨会话对齐**的录音："
            "同一音源标识下，【被测设备】与【对比设备】分别在两次会话中单独外放，经同一麦克风链路分时采集。"
            "请仅依据听感完成被测相对对比的**五维整数量化**，并输出规定 JSON。"
        )

    _mm_block = (
        "【多模态与附件事实】本请求已通过 Dify 接口在消息中附带 **audio 类型本地文件**（local_file + upload_file_id），"
        "每路均为已上传的 **完整 WAV 音频数据**，并非仅有文件名。你必须使用应用提供的 **音频理解能力** 基于波形/听感完成比较。"
        "**禁止**以「仅有文件名」「无法播放/试听」「无法实际试听」「无法直接试听」「无法试听附件」「无法获取音频流」"
        "「信息不足以主观对比」「无法解析音频」等为由拒绝输出下方规定 JSON 或五维全 0 占位；"
        "若电平极低或接近静音，仍须输出合法 JSON，并在「专业点评」中说明依据。"
        "\n\n"
    )

    prefix = f"{body}\n\n" if body else ""
    return f"""{prefix}{cross_block}
{intro_line}

{_mm_block}附件与槽位对应（请严格按顺序解读）：
{mapping_lines}

**被测 / 对比与附件序号**：以说明为准；若无额外说明，则约定 **第 {dut_n} 个附件 = 被测设备**，**第 {ref_n} 个附件 = 对比设备**。请仅基于这两路（在有多路时以被测相对对比设备为主）完成下列评分。

评分原则（必须遵守）：
- 仅比较【被测设备】相对【对比设备】在同一音源下的内置喇叭表现，依据主观听感。
- 下列五个维度**必须各自单独给出一个整数分**，分值**只能是** -3、-2、-1、0、1、2、3 之一：
  **正分**表示被测优于对比；**负分**表示被测劣于对比；**0** 表示两者相当。
- **禁止**输出小数、禁止输出区间（如「1～2」「约1」）、禁止合并维度给分。
- 不得用其它数字（如 4、-4、10）。

**必须只输出一个 JSON 对象**，不要 Markdown 代码围栏，不要其它文字。JSON Schema（键名与顺序保持一致）：
{{
  "音源": {stim_json},
  "声音响度": <int, 仅 -3,-2,-1,0,1,2,3>,
  "人声清晰度": <int, 仅 -3,-2,-1,0,1,2,3>,
  "听感舒适度": <int, 仅 -3,-2,-1,0,1,2,3>,
  "失真与噪声": <int, 仅 -3,-2,-1,0,1,2,3>,
  "频响平衡": <int, 仅 -3,-2,-1,0,1,2,3>,
  "对比总结": "<中文，80字内，概括被测相对对比的整体听感结论>",
  "专业点评": "<中文，不超过200字，可含改进建议>"
}}

**流式输出要求**：请将上述 JSON **完整**作为助手回复的正文输出（确保内容出现在接口可识别的 answer/文本流中），勿留空；涉及《月光小夜曲》《赤壁怀古》等长曲名时仍须输出合法 JSON。
"""


class DifyClient:
    """Dify 对话与文件上传客户端（密钥优先读环境变量 DIFY_API_KEY）。"""

    def __init__(self, log: Optional[Callable[[str], None]] = None) -> None:
        try:
            from config import (  # type: ignore
                DIFY_API_KEY,
                DIFY_API_URL,
                DIFY_FILE_UPLOAD_URL,
                DIFY_USER,
            )

            default_key, default_url, default_upload, default_user = (
                DIFY_API_KEY,
                DIFY_API_URL,
                DIFY_FILE_UPLOAD_URL,
                DIFY_USER,
            )
        except Exception as exc:
            print(f"[Dify] 读取 config 中的 Dify 默认配置失败，将仅使用环境变量/空默认值：{exc}", file=sys.stderr)
            default_key = ""
            default_url = "https://dify.cvte.com/v1/chat-messages"
            default_upload = "https://dify.cvte.com/v1/files/upload"
            default_user = ""

        self.api_key = os.environ.get("DIFY_API_KEY", default_key)
        self.api_url = os.environ.get("DIFY_API_URL", default_url)
        self.file_upload_url = os.environ.get("DIFY_FILE_UPLOAD_URL", default_upload)
        self.user = os.environ.get("DIFY_USER", default_user)
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        self._user_log = log

    def _emit(self, msg: str) -> None:
        """子进程/Web UI：优先写入 scoring 传入的 log；否则打印 stdout。"""
        if self._user_log:
            self._user_log(msg)
        else:
            print(msg, flush=True)

    def _upload_file_multipart_with_retry(
        self,
        path: Path,
        mime_type: str,
        log_label: str,
    ) -> dict:
        """multipart 上传，失败时最多重试 3 次（间隔 5s），成功返回 JSON dict（含 id）。"""
        if not path.exists():
            raise FileNotFoundError(f"上传文件不存在：{path}")
        last_err: Optional[str] = None
        for attempt in range(DIFY_RETRY_MAX_ATTEMPTS):
            if attempt > 0:
                self._emit(
                    f"[Dify] {log_label}：重试第{attempt}次（上次：{last_err}）…"
                )
                time.sleep(DIFY_RETRY_INTERVAL_SEC)
            else:
                self._emit(f"[Dify] {log_label}：准备上传附件…")
                _live_scoring_detail(f"{log_label}：正在上传附件…")
            try:
                _tup = _file_upload_timeout()
                try:
                    _sz = path.stat().st_size if path.is_file() else -1
                except OSError:
                    _sz = -1
                self._emit(
                    f"[Dify] {log_label}：POST multipart（约 {_sz} bytes，连接/读超时 {_tup[0]:.0f}s / {_tup[1]:.0f}s）…"
                )
                with open(path, "rb") as f:
                    files = {"file": (path.name, f, mime_type)}
                    data = {"user": self.user}
                    headers = {"Authorization": f"Bearer {self.api_key}"}
                    response = requests.post(
                        self.file_upload_url,
                        headers=headers,
                        files=files,
                        data=data,
                        timeout=_tup,
                    )
                if response.status_code == 201:
                    result = response.json()
                    # #region agent log
                    _agent_dbg_upload(
                        "H2",
                        "difyclient._upload_file_multipart_with_retry:201",
                        "upload_ok",
                        {
                            "log_label": log_label[:120],
                            "path_name": path.name[:200],
                            "file_bytes": path.stat().st_size if path.is_file() else -1,
                            "file_id": str(result.get("id", ""))[:64],
                        },
                    )
                    # #endregion
                    self._emit(f"[Dify] {log_label}：上传成功")
                    _live_scoring_detail(f"{log_label}：附件已上传")
                    return result
                last_err = f"HTTP {response.status_code}: {response.text[:1200]}"
                # #region agent log
                _agent_dbg_upload(
                    "H2",
                    "difyclient._upload_file_multipart_with_retry:http_error",
                    "non_201",
                    {
                        "status": response.status_code,
                        "log_label": log_label[:120],
                        "body_snip": response.text[:500],
                    },
                )
                # #endregion
            except requests.RequestException as exc:
                last_err = f"{type(exc).__name__}: {exc}"
                # #region agent log
                _agent_dbg_upload(
                    "H3",
                    "difyclient._upload_file_multipart_with_retry:net",
                    "request_exception",
                    {"log_label": log_label[:120], "err": str(exc)[:500]},
                )
                # #endregion
        self._emit(f"[Dify] {log_label}：上传失败（已重试3次）— {last_err}")
        # #region agent log
        _agent_dbg_upload(
            "H2",
            "difyclient._upload_file_multipart_with_retry:exhausted",
            "all_attempts_failed",
            {"log_label": log_label[:120], "last_err": (last_err or "")[:800]},
        )
        # #endregion
        raise RuntimeError(
            f"Dify 文件上传失败（{log_label}），共 {DIFY_RETRY_MAX_ATTEMPTS} 次尝试仍失败：{last_err}"
        )

    def _post_chat_messages_stream_with_retry(
        self,
        data: dict,
        log_label: str,
        timeout: int | float | tuple[float, float],
    ) -> requests.Response:
        """chat-messages 流式 POST，失败时重试；返回 status_code==200 的 Response。"""
        last_err: Optional[str] = None
        payload = json.dumps(data)
        http_timeout = _chat_request_timeout(float(timeout)) if not isinstance(
            timeout, tuple
        ) else timeout
        for attempt in range(DIFY_RETRY_MAX_ATTEMPTS):
            if attempt > 0:
                self._emit(
                    f"[Dify] {log_label}：重试第{attempt}次（上次：{last_err}）…"
                )
                time.sleep(DIFY_RETRY_INTERVAL_SEC)
            else:
                self._emit(f"[Dify] {log_label}：评分请求中（等待 Dify/模型）…")
                _live_scoring_detail(f"{log_label}：正在请求 Dify 评分…")
            try:
                response = requests.post(
                    self.api_url,
                    headers=self.headers,
                    data=payload,
                    stream=True,
                    timeout=http_timeout,
                )
                if response.status_code == 200:
                    self._emit(f"[Dify] {log_label}：已连接，正在接收模型输出…")
                    _live_scoring_detail(f"{log_label}：正在接收模型流式回复…")
                    return response
                last_err = f"HTTP {response.status_code}: {response.text[:1200]}"
            except requests.RequestException as exc:
                last_err = f"{type(exc).__name__}: {exc}"
        self._emit(f"[Dify] {log_label}：请求失败（已重试3次）— {last_err}")
        raise RuntimeError(
            f"Dify 评分请求失败（{log_label}），共 {DIFY_RETRY_MAX_ATTEMPTS} 次尝试仍失败：{last_err}"
        )

    @staticmethod
    def _detect_mime_and_file_type(path: Path) -> tuple[str, str]:
        """
        检测文件 MIME 与 Dify 文件分类，避免回落到 application/octet-stream。
        返回: (mime_type, file_type) 其中 file_type 为 image/document。
        """
        extension = path.suffix.lower()
        custom_map = {
            # Vertex/Gemini 对 text/csv 兼容性不稳定，统一按 text/plain 处理更稳
            ".csv": "text/plain",
            ".tsv": "text/tab-separated-values",
            ".md": "text/markdown",
            ".json": "application/json",
            ".yaml": "text/yaml",
            ".yml": "text/yaml",
        }
        image_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}

        mime_type = custom_map.get(extension) or mimetypes.guess_type(path.name)[0]
        if not mime_type or mime_type == "application/octet-stream":
            # Vertex/Gemini 不接受 octet-stream，兜底改为 text/plain。
            mime_type = "text/plain"

        # 再做一次白名单兜底，避免后端把不支持的 mime 透传给模型
        allowed_exact = {
            "text/plain",
            "application/pdf",
            "application/json",
            "text/markdown",
            "text/yaml",
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-powerpoint",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        }
        if not (mime_type.startswith("image/") or mime_type in allowed_exact):
            mime_type = "text/plain"

        if extension in image_exts or mime_type.startswith("image/"):
            file_type = "image"
        else:
            file_type = "document"

        return mime_type, file_type

    def _parse_sse_line(self, line: str) -> Optional[dict]:
        """
        解析单行 SSE。兼容 ``data:`` 后无空格、以及 ``: ping`` 心跳注释行。
        """
        line = line.strip()
        if not line:
            return None
        if line.startswith(":"):
            return None
        if line.startswith("data:"):
            payload = line[5:].lstrip()
        else:
            payload = line
        if payload in ("", "[DONE]"):
            return None
        try:
            obj = json.loads(payload)
        except json.JSONDecodeError:
            return None
        if isinstance(obj, dict) and obj.get("event") == "error":
            return obj
        return obj if isinstance(obj, dict) else None

    @staticmethod
    def _looks_like_usage_only_json_text(s: str) -> bool:
        """
        Dify/工作流末包常见「用量/计费」JSON，无评分键；若当作正文会污染聚合并导致 Web 全 0。
        """
        t = (s or "").strip()
        if not t.startswith("{"):
            return False
        try:
            d = json.loads(t)
        except json.JSONDecodeError:
            return False
        if not isinstance(d, dict):
            return False
        keys = set(d.keys())
        if "prompt_tokens" in keys and "total_tokens" in keys:
            return True
        if "latency" in keys and "completion_unit_price" in keys and "currency" in keys:
            return True
        return False

    @staticmethod
    def _extract_answer_text_from_event(chunk: dict) -> Optional[str]:
        """
        从单条 SSE JSON 中取模型正文。优先 ``answer``，兼容 Doubao / 工作流等嵌套在 ``data``、
        ``metadata``、``outputs`` 中的 text/content（部分模型不填顶层 answer）。
        """
        if not isinstance(chunk, dict):
            return None

        def _emit(s: Optional[str]) -> Optional[str]:
            if not isinstance(s, str):
                return None
            t = s.strip()
            if not t or DifyClient._looks_like_usage_only_json_text(t):
                return None
            return t

        def _take_str(d: dict, keys: tuple[str, ...]) -> Optional[str]:
            for k in keys:
                v = d.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip()
            return None

        t = _take_str(chunk, ("answer", "text", "content", "output", "result"))
        if t:
            e = _emit(t)
            if e:
                return e

        top_outs = chunk.get("outputs")
        if isinstance(top_outs, dict):
            t = _take_str(top_outs, ("answer", "text", "content", "output", "result"))
            if t:
                e = _emit(t)
                if e:
                    return e
            nested = DifyClient._scoring_dict_to_answer_text(top_outs)
            if nested:
                e = _emit(nested)
                if e:
                    return e

        data = chunk.get("data")
        if isinstance(data, dict):
            t = _take_str(data, ("answer", "text", "content", "output", "result"))
            if t:
                e = _emit(t)
                if e:
                    return e
            outs = data.get("outputs")
            if isinstance(outs, dict):
                t = _take_str(outs, ("answer", "text", "content", "output", "result"))
                if t:
                    e = _emit(t)
                    if e:
                        return e
                nested = DifyClient._scoring_dict_to_answer_text(outs)
                if nested:
                    e = _emit(nested)
                    if e:
                        return e
            if isinstance(outs, list):
                for item in outs:
                    if isinstance(item, dict):
                        t = _take_str(item, ("answer", "text", "content", "output", "result"))
                        if t:
                            e = _emit(t)
                            if e:
                                return e
                        nested = DifyClient._scoring_dict_to_answer_text(item)
                        if nested:
                            e = _emit(nested)
                            if e:
                                return e

            nested_data = DifyClient._scoring_dict_to_answer_text(data)
            if nested_data:
                e = _emit(nested_data)
                if e:
                    return e

        for meta in (chunk.get("metadata"), chunk.get("message_metadata")):
            if isinstance(meta, dict):
                t = _take_str(meta, ("answer", "content", "text"))
                if t:
                    e = _emit(t)
                    if e:
                        return e
                nested_m = DifyClient._scoring_dict_to_answer_text(meta)
                if nested_m:
                    e = _emit(nested_m)
                    if e:
                        return e

        msg = chunk.get("message")
        if isinstance(msg, dict):
            c = msg.get("content")
            if isinstance(c, str) and c.strip():
                e = _emit(c.strip())
                if e:
                    return e
            if isinstance(c, list):
                parts: list[str] = []
                for block in c:
                    if isinstance(block, dict) and block.get("type") == "text":
                        tx = block.get("text")
                        if isinstance(tx, str) and tx:
                            parts.append(tx)
                    elif isinstance(block, str):
                        parts.append(block)
                merged = "".join(parts).strip()
                if merged:
                    e = _emit(merged)
                    if e:
                        return e

        ch = chunk.get("choices")
        if isinstance(ch, list) and ch:
            z0 = ch[0]
            if isinstance(z0, dict):
                dlt = z0.get("delta") or z0.get("message")
                if isinstance(dlt, dict):
                    tc = dlt.get("content")
                    if isinstance(tc, str) and tc.strip():
                        e = _emit(tc.strip())
                        if e:
                            return e

        deep = DifyClient._deep_scan_strings_for_answer(chunk)
        if deep:
            return deep

        return None

    @staticmethod
    def _parsed_dict_has_scoring_dims(d: dict) -> bool:
        """与 scoring 侧一致：须含至少一个五维键，避免仅 ``音源`` 等短片段误匹配。"""
        return any(
            k in d
            for k in (
                "声音响度",
                "人声清晰度",
                "听感舒适度",
                "失真与噪声",
                "频响平衡",
            )
        )

    @staticmethod
    def _find_first_dict_with_scoring_dims(
        node: Any, depth: int = 0, max_depth: int = 14
    ) -> Optional[dict]:
        """
        部分 Dify 工作流（尤其 Gemini）把五维分放在 ``outputs`` 的结构体里，不写顶层 ``answer``。
        在任意嵌套对象中找出**首个**含五维键的 dict（通常为可直接 JSON 化的评分对象）。
        """
        if depth > max_depth or node is None:
            return None
        if isinstance(node, dict):
            if DifyClient._parsed_dict_has_scoring_dims(node):
                return node
            for v in node.values():
                r = DifyClient._find_first_dict_with_scoring_dims(v, depth + 1, max_depth)
                if r is not None:
                    return r
        elif isinstance(node, list):
            for it in node:
                r = DifyClient._find_first_dict_with_scoring_dims(it, depth + 1, max_depth)
                if r is not None:
                    return r
        return None

    @staticmethod
    def _scoring_dict_to_answer_text(container: dict) -> Optional[str]:
        """若 container 自身或其子树中含五维评分 dict，则序列化为与 ``answer`` 等价的正文。"""
        sd = DifyClient._find_first_dict_with_scoring_dims(container)
        if not sd:
            return None
        try:
            s = json.dumps(sd, ensure_ascii=False)
        except (TypeError, ValueError):
            return None
        t = (s or "").strip()
        if not t or DifyClient._looks_like_usage_only_json_text(t):
            return None
        return t

    @staticmethod
    def _deep_scan_strings_for_answer(
        node: Any, depth: int = 0, max_depth: int = 16
    ) -> Optional[str]:
        """
        工作流 ``node_finished`` / ``workflow_finished`` 等事件中，正文常在 ``data.outputs`` 的
        任意键下且为 JSON 字符串；从整棵子树中找出可解析且含五维键的 JSON 文本。
        """
        if depth > max_depth:
            return None
        if isinstance(node, str):
            t = node.strip()
            if len(t) < 10 or not t.startswith("{"):
                return None
            if DifyClient._looks_like_usage_only_json_text(t):
                return None
            if "声音响度" not in t and "人声清晰度" not in t:
                return None
            try:
                d = json.loads(t)
            except json.JSONDecodeError:
                return None
            if isinstance(d, dict) and DifyClient._parsed_dict_has_scoring_dims(d):
                return t
            return None
        if isinstance(node, dict):
            outs = node.get("outputs")
            if isinstance(outs, dict):
                for v in outs.values():
                    r = DifyClient._deep_scan_strings_for_answer(v, depth + 1, max_depth)
                    if r:
                        return r
            for v in node.values():
                r = DifyClient._deep_scan_strings_for_answer(v, depth + 1, max_depth)
                if r:
                    return r
        if isinstance(node, list):
            for it in node:
                r = DifyClient._deep_scan_strings_for_answer(it, depth + 1, max_depth)
                if r:
                    return r
        return None

    @staticmethod
    def _normalize_sse_physical_lines(lines: list[str]) -> list[str]:
        """同一物理行内紧挨多个 ``data:{...}`` 时拆成多段，避免整包只被解析一次。"""
        out: list[str] = []
        for raw in lines:
            s = raw.strip()
            if not s:
                continue
            if s.count("data:") <= 1:
                out.append(raw)
                continue
            idx = 0
            while idx < len(s):
                hit = s.find("data:", idx)
                if hit < 0:
                    break
                nxt = s.find("data:", hit + 5)
                piece = (s[hit:nxt] if nxt >= 0 else s[hit:]).strip()
                if piece:
                    out.append(piece)
                if nxt < 0:
                    break
                idx = nxt
        return out

    def _fallback_extract_scoring_text_from_sse_lines(self, lines: list[str]) -> str:
        """
        当增量 answer 全空时，从原始 SSE 拼接串中抠出 JSON 评分块（适配长音源名、弯引号等导致
        仅出现在末包或非标准字段的情况）。
        """
        blob = "\n".join(lines)
        # 子串须含至少一个五维键名；勿仅用「音源」——易命中不完整 JSON（日志中 fb_len≈270 且解析无五维）。
        dim_markers = (
            "声音响度",
            "人声清晰度",
            "听感舒适度",
            "失真与噪声",
            "频响平衡",
        )

        def _candidate_ok(s: str) -> bool:
            return any(k in s for k in dim_markers)

        candidates: list[str] = []
        for pattern in (
            r"\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}",
            r"\{[\s\S]*?\}",
        ):
            for m in re.finditer(pattern, blob):
                s = m.group(0).strip()
                if not _candidate_ok(s):
                    continue
                try:
                    parsed = json.loads(s)
                except json.JSONDecodeError:
                    continue
                if isinstance(parsed, dict) and self._parsed_dict_has_scoring_dims(parsed):
                    candidates.append(s)
        if candidates:
            return max(candidates, key=len)

        start = blob.rfind("{")
        if start < 0:
            return ""
        depth = 0
        for i in range(start, len(blob)):
            if blob[i] == "{":
                depth += 1
            elif blob[i] == "}":
                depth -= 1
                if depth == 0:
                    cand = blob[start : i + 1]
                    if not _candidate_ok(cand):
                        return ""
                    try:
                        parsed = json.loads(cand)
                    except json.JSONDecodeError:
                        break
                    if isinstance(parsed, dict) and self._parsed_dict_has_scoring_dims(
                        parsed
                    ):
                        return cand
                    break
        return ""

    def _aggregate_answer_from_sse_lines(self, lines: list[str]) -> str:
        """与 collect_stream_answer 相同合并策略，但使用扩展字段提取。"""
        cumulative: Optional[str] = None
        delta_parts: list[str] = []

        for line in lines:
            chunk = self._parse_sse_line(line)
            if not chunk:
                continue
            if chunk.get("event") == "error":
                continue
            ans = self._extract_answer_text_from_event(chunk)
            if not ans:
                continue

            if cumulative is None:
                cumulative = ans
                delta_parts = [ans]
                continue

            if ans.startswith(cumulative):
                cumulative = ans
                delta_parts = [ans]
            elif cumulative.startswith(ans):
                pass
            else:
                delta_parts.append(ans)
                cumulative = "".join(delta_parts)

        return (cumulative or "").strip()

    def collect_stream_answer(
        self, response: requests.Response, *, log_label: str = ""
    ) -> str:
        """从 Dify 流式响应中取出完整回答（兼容 answer / data 嵌套及末段 JSON 兜底）。"""
        lines: list[str] = []
        wall = _stream_collect_wall_sec()
        deadline = time.monotonic() + wall
        progress_iv = _stream_progress_interval_sec()
        last_progress = time.monotonic()
        line_iter = response.iter_lines(decode_unicode=True)
        _lbl = (log_label or "流式").strip()

        while True:
            now = time.monotonic()
            if now >= deadline:
                self._emit(
                    f"[Dify] {_lbl}：流式读取总时长已超过 {wall:.0f}s，停止等待。"
                )
                _live_scoring_detail(
                    f"{_lbl}：等待模型超时（>{wall:.0f}s），将尝试其它方式或失败"
                )
                break
            if now - last_progress >= progress_iv:
                elapsed = wall - (deadline - now)
                self._emit(
                    f"[Dify] {_lbl}：仍在等待模型输出（已约 {elapsed:.0f}s / {wall:.0f}s）…"
                )
                _live_scoring_detail(
                    f"{_lbl}：仍在等待模型输出（约 {elapsed:.0f}s）…"
                )
                last_progress = now
            try:
                raw = next(line_iter)
            except StopIteration:
                break
            except requests.exceptions.ReadTimeout:
                self._emit(
                    f"[Dify] {_lbl}：相邻数据块等待超时，停止流式读取。"
                )
                _live_scoring_detail(f"{_lbl}：模型长时间无数据返回")
                break
            if raw is None:
                continue
            s = raw if isinstance(raw, str) else raw.decode("utf-8", errors="replace")
            if s.strip():
                lines.append(s)
                last_progress = time.monotonic()

        stream_err: Optional[str] = None
        for ln in lines:
            ch = self._parse_sse_line(ln)
            if ch and ch.get("event") == "error":
                stream_err = str(ch.get("message") or ch.get("code") or ch)[:500]
                self._emit(f"[Dify] 流式错误事件：{stream_err}")
                _live_scoring_detail(
                    f"{_lbl}：Dify 返回错误 — {stream_err[:120]}"
                )

        lines = self._normalize_sse_physical_lines(lines)
        agg = self._aggregate_answer_from_sse_lines(lines)
        fb = (
            self._fallback_extract_scoring_text_from_sse_lines(lines) if not agg else ""
        )
        text = agg or fb
        if not (text or "").strip():
            for ln in lines:
                ch = self._parse_sse_line(ln)
                if not ch:
                    continue
                sd = DifyClient._find_first_dict_with_scoring_dims(ch)
                if sd:
                    try:
                        text = json.dumps(sd, ensure_ascii=False)
                    except (TypeError, ValueError):
                        pass
                    if text:
                        self._emit(
                            f"[Dify] {_lbl}：从 SSE 嵌套结构中提取到评分 JSON（{len(text)} 字符）。"
                        )
                        break
        if not (text or "").strip() and lines:
            ev_names: list[str] = []
            for ln in lines:
                ch = self._parse_sse_line(ln)
                if ch and ch.get("event"):
                    ev_names.append(str(ch.get("event")))
            tail = "、".join(ev_names[-6:]) if ev_names else "（无 event 字段）"
            self._emit(
                f"[Dify] {_lbl}：共收到 {len(lines)} 行 SSE，但未解析出模型正文；"
                f"末段事件：{tail}。"
                + (" 流式错误：" + stream_err[:200] if stream_err else "")
            )
        # #region agent log
        _agent_dbg_upload(
            "H1",
            "difyclient.py:collect_stream_answer",
            "stream_aggregate",
            {
                "n_sse_lines": len(lines),
                "agg_len": len(agg),
                "fb_len": len(fb),
                "out_len": len(text or ""),
                "used_fallback": bool(fb),
                "stream_err": (stream_err or "")[:200],
            },
        )
        # #endregion
        return text or ""

    def _post_and_collect_stream_answer(
        self, data: dict, log_label: str, timeout: int | float
    ) -> str:
        """
        发起流式对话并解析正文；若正文为空则**再流式重试 1 次**；仍空则 **blocking** 同步拉取
        （适配工作流仅把结果放在 ``node_finished.outputs``、或网关缓冲导致 SSE 无 ``answer`` 等情形）。
        """
        stream_attempts = _empty_body_stream_attempts()
        empty_hint = (
            "未从流式事件中解析到模型正文（部分模型/工作流可能不写 answer、或仅末包有 outputs）；"
            f"1 秒后将自动重试流式（最多 {stream_attempts} 次）；"
            "若仍失败将再尝试 blocking 模式（可用 DIFY_SKIP_BLOCKING_FALLBACK=1 跳过）。"
        )
        http_timeout = _chat_request_timeout(float(timeout))
        for attempt in range(stream_attempts):
            if attempt > 0:
                self._emit(f"[Dify] {log_label}：{empty_hint}")
                time.sleep(1.0)
            label = log_label if attempt == 0 else f"{log_label}·正文补采"
            resp = self._post_chat_messages_stream_with_retry(data, label, http_timeout)
            try:
                text = self.collect_stream_answer(resp, log_label=log_label)
            finally:
                try:
                    resp.close()
                except Exception:
                    pass
            # #region agent log
            _agent_dbg_upload(
                "H2",
                "difyclient.py:_post_and_collect_stream_answer",
                "collect_attempt",
                {
                    "log_label": log_label[:80],
                    "attempt": attempt,
                    "text_len": len(text or ""),
                    "text_nonempty": bool((text or "").strip()),
                },
            )
            # #endregion
            if text and text.strip():
                if attempt > 0:
                    self._emit(f"[Dify] {log_label}：正文补采成功，已得到模型输出。")
                return text

        if _env_truthy("DIFY_SKIP_BLOCKING_FALLBACK"):
            self._emit(
                f"[Dify] {log_label}：流式 {stream_attempts} 次均无正文；"
                "已设置 DIFY_SKIP_BLOCKING_FALLBACK，跳过 blocking。"
            )
            _live_scoring_detail(f"{log_label}：Dify 未返回可用正文")
            return ""

        btxt = self._post_chat_messages_blocking_collect(
            data, log_label, http_timeout
        )
        if btxt and btxt.strip():
            return btxt.strip()

        self._emit(
            f"[Dify] {log_label}：流式 {stream_attempts} 次 + blocking 兜底后仍无可用正文；"
            "请检查 Dify 工作流末节点是否将回复写入对话 answer、或是否仅在工作流变量中输出。"
        )
        _live_scoring_detail(f"{log_label}：Dify 未返回可用正文，请检查工作流或模型")
        return ""

    def _extract_blocking_answer_from_json(self, body: dict) -> str:
        """blocking 模式返回的 JSON 中取助手正文。"""
        if not isinstance(body, dict):
            return ""
        ans = body.get("answer")
        if isinstance(ans, str) and ans.strip():
            if not DifyClient._looks_like_usage_only_json_text(ans):
                return ans.strip()
        for probe in (body.get("data"), body.get("outputs"), body.get("metadata"), body):
            if isinstance(probe, dict):
                sc = DifyClient._scoring_dict_to_answer_text(probe)
                if sc:
                    return sc.strip()
        deep = DifyClient._deep_scan_strings_for_answer(body)
        if deep:
            return deep
        return ""

    def _post_chat_messages_blocking_collect(
        self,
        data: dict,
        log_label: str,
        timeout: int | float | tuple[float, float],
    ) -> str:
        """
        流式无法解析正文时，用 ``response_mode=blocking`` 再取一次完整 JSON（与流式共用 HTTP 重试策略）。
        """
        payload = dict(data)
        payload["response_mode"] = "blocking"
        body = json.dumps(payload)
        last_err: Optional[str] = None
        http_timeout = timeout if isinstance(timeout, tuple) else _chat_request_timeout(
            float(timeout)
        )
        max_attempts = _blocking_max_attempts()
        for attempt in range(max_attempts):
            if attempt > 0:
                self._emit(
                    f"[Dify] {log_label}·blocking：重试第{attempt}次（上次：{last_err}）…"
                )
                time.sleep(DIFY_RETRY_INTERVAL_SEC)
            else:
                self._emit(
                    f"[Dify] {log_label}·blocking：流式无正文，改用同步 blocking 拉取完整回复…"
                )
                _live_scoring_detail(f"{log_label}：流式无正文，尝试 blocking 同步拉取…")
            try:
                r = requests.post(
                    self.api_url,
                    headers=self.headers,
                    data=body,
                    stream=False,
                    timeout=http_timeout,
                )
                if r.status_code != 200:
                    last_err = f"HTTP {r.status_code}: {r.text[:800]}"
                    continue
                try:
                    j = r.json()
                except json.JSONDecodeError as exc:
                    last_err = f"JSON解析失败: {exc}"
                    continue
                if not isinstance(j, dict):
                    last_err = "blocking 响应非 JSON 对象"
                    continue
                if j.get("event") == "error" or j.get("code"):
                    last_err = str(j.get("message") or j.get("code") or j)[:800]
                    continue
                text = self._extract_blocking_answer_from_json(j)
                if text.strip():
                    self._emit(
                        f"[Dify] {log_label}·blocking：已取到正文（{len(text)} 字符）。"
                    )
                    return text
                last_err = "blocking 响应中 answer/outputs 均无可用正文"
            except requests.RequestException as exc:
                last_err = f"{type(exc).__name__}: {exc}"
        self._emit(f"[Dify] {log_label}·blocking：失败 — {last_err}")
        return ""

    @staticmethod
    def _read_text_file_for_prompt(path: Path, max_chars: int = 60000) -> str:
        """读取文本文件用于拼接提问，超长会截断。"""
        text = path.read_text(encoding="utf-8", errors="replace")
        if len(text) > max_chars:
            return text[:max_chars] + "\n...(内容已截断)"
        return text

    @staticmethod
    def _read_xlsx_for_prompt(
        path: Path, max_rows_per_sheet: int = 200, max_chars: int = 60000
    ) -> str:
        """
        读取 xlsx 内容并转为可提问文本（按 sheet 输出，行内用 TAB 分隔）。
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            blocks: list[str] = []
            for ws in wb.worksheets:
                blocks.append(f"# Sheet: {ws.title}")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    if row_count > max_rows_per_sheet:
                        blocks.append("...(该工作表内容已截断)")
                        break
                    values = ["" if v is None else str(v) for v in row]
                    line = "\t".join(values).rstrip("\t")
                    if line:
                        blocks.append(line)
                blocks.append("")

            text = "\n".join(blocks).strip()
            if len(text) > max_chars:
                return text[:max_chars] + "\n...(内容已截断)"
            return text
        finally:
            wb.close()

    def _upload_audio_for_chat(self, path: Path, upload_mime: str) -> dict:
        """
        将音频规范为 **48kHz / 单声道 / 16-bit PCM WAV** 后，再复制为纯 ASCII 短文件名上传。

        制式仅含采样率/声道/位深转换（``audio_llm_normalize``），**不进行**响度/LUFS/峰值归一化，
        以保留实录相对电平。原始路径 ``path`` 处文件不修改；仅临时文件参与上传。
        部分 Dify 应用在 chat-messages 阶段会校验附件名；含中文或弯引号（如 '’‘）的长名会报 invalid_param。

        ``upload_mime`` 仅保留兼容旧调用；实际上传内容恒为 WAV，故 MIME 固定为 ``audio/wav``。
        """
        del upload_mime  # 上传前已统一为 WAV
        tdir = Path(tempfile.mkdtemp(prefix="dify_audio_"))
        staged = tdir / f"dify_{uuid.uuid4().hex[:16]}.wav"
        try:
            # #region agent log
            try:
                _ob = path.stat().st_size if path.is_file() else -1
            except OSError:
                _ob = -1
            _agent_dbg_upload(
                "H5",
                "difyclient._upload_audio_for_chat:entry",
                "before_normalize",
                {"orig": str(path.name)[:200], "orig_bytes": _ob},
            )
            # #endregion
            from audio_llm_normalize import write_normalized_wav_for_upload

            _cap = dify_upload_max_audio_seconds()
            _cap_label = (
                f"上传最长 {_cap:.0f}s（超出则仅取开头）"
                if _cap is not None
                else "上传不截断"
            )
            self._emit(f"[Dify] 音频 {path.name}：正在规范化为 48kHz WAV（{_cap_label}）…")
            _live_scoring_detail(f"音频 {path.name[:48]}：正在规范化…")
            try:
                _norm_meta = write_normalized_wav_for_upload(
                    path, staged, max_duration_sec=_cap
                )
            except Exception as _norm_exc:
                # #region agent log
                _agent_dbg_upload(
                    "H1",
                    "difyclient._upload_audio_for_chat:normalize_exc",
                    "write_normalized_failed",
                    {
                        "exc_type": type(_norm_exc).__name__,
                        "exc_msg": str(_norm_exc)[:600],
                    },
                )
                # #endregion
                raise
            if _norm_meta.get("trimmed"):
                _din = float(_norm_meta.get("duration_in_sec") or 0.0)
                _dout = float(_norm_meta.get("duration_out_sec") or 0.0)
                self._emit(
                    f"[Dify] 音频 {path.name}：实录约 {_din:.1f}s，"
                    f"上传截断为前 {_dout:.1f}s（上限 {_cap:.0f}s）"
                )
                _live_scoring_detail(
                    f"音频 {path.name[:32]}：上传前截断 {_din:.0f}s→{_dout:.0f}s"
                )
            # #region agent log
            try:
                _sb = staged.stat().st_size if staged.is_file() else 0
            except OSError:
                _sb = 0
            _agent_dbg_upload(
                "H4",
                "difyclient._upload_audio_for_chat:after_normalize",
                "staged_ready",
                {
                    "staged_bytes": _sb,
                    "staged_name": staged.name[:120],
                    "trimmed": bool(_norm_meta.get("trimmed")),
                    "duration_in_sec": _norm_meta.get("duration_in_sec"),
                    "duration_out_sec": _norm_meta.get("duration_out_sec"),
                },
            )
            # #endregion
            return self.upload_file_generic(
                str(staged),
                mime_type="audio/wav",
                upload_label=f"音频 {path.name}",
            )
        finally:
            shutil.rmtree(tdir, ignore_errors=True)

    def upload_file(self, file_path: str) -> dict:
        """
        上传文件到 Dify
        Args:
            file_path: 本地文件路径
        Returns:
            dict: 包含文件信息的字典，如 {"id": "xxx", "name": "filename.jpg"}
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"上传文件不存在：{file_path}")
        mime_type, _ = self._detect_mime_and_file_type(path)
        return self.upload_file_generic(
            file_path, mime_type, upload_label=f"文件 {path.name}"
        )

    def rum(self, problem):
        data = {
            "inputs": {},
            "query": problem,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": self.user,
        }
        text = self._post_and_collect_stream_answer(data, "对话", 120)
        if text:
            print(text)
            return text
        return None

    def chat_with_image(self, image_path: str, query: str = "请描述这张图片"):
        """
        使用图片进行对话

        Args:
            image_path: 本地图片路径
            query: 问题
        """
        # 第一步：上传图片
        upload_result = self.upload_file(image_path)

        file_id = upload_result.get("id")
        file_name = upload_result.get("name", Path(image_path).name)

        if not file_id:
            raise RuntimeError("Dify 上传返回中缺少文件 id")

        # 第二步：使用上传的文件 ID 进行对话
        data = {
            "inputs": {},
            "query": query,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": self.user,
            "files": [
                {
                    "type": "image",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id,
                    "filename": file_name
                }
            ]
        }

        text = self._post_and_collect_stream_answer(data, "图片对话", 120)
        if text:
            return text
        return None

    def analyze_file(self, file_path: str, query: str = "请分析这个文件的内容"):
        """
        上传文件并针对文件内容进行提问

        Args:
            file_path: 本地文件路径（支持图片、文档等多种格式）
            query: 针对文件的问题

        Returns:
            str: AI 的回答内容
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{file_path}")

        # 自动检测 MIME 与文件分类，避免 octet-stream
        mime_type, file_type = self._detect_mime_and_file_type(path)

        # 文本类文件直接以内文提问，绕开附件链路的模型参数限制
        text_like_exts = {".csv", ".tsv", ".txt", ".md", ".json", ".yaml", ".yml"}
        if path.suffix.lower() in text_like_exts:
            file_text = self._read_text_file_for_prompt(path)
            prompt = (
                f"{query}\n\n"
                f"文件名：{path.name}\n"
                f"以下是文件内容：\n"
                f"``text\n{file_text}\n```"
            )
            return self.rum(prompt)

        # xlsx 先提取表格文本再提问，避免模型不支持该 mime
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            file_text = self._read_xlsx_for_prompt(path)
            if not file_text:
                print("无法解析 xlsx：请安装 openpyxl 或先另存为 csv。")
                return None
            prompt = (
                f"{query}\n\n"
                f"文件名：{path.name}\n"
                f"以下是工作簿提取内容：\n"
                f"``text\n{file_text}\n```"
            )
            return self.rum(prompt)

        # 第一步：上传文件
        upload_result = self.upload_file_generic(
            file_path, mime_type, upload_label=f"文件 {path.name}"
        )

        file_id = upload_result.get("id")
        file_name = upload_result.get("name", path.name)

        if not file_id:
            raise RuntimeError("Dify 上传返回中缺少文件 id")

        # 第二步：使用上传的文件 ID 进行对话
        data = {
            "inputs": {},
            "query": query,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": self.user,
            "files": [
                {
                    "type": file_type,
                    "transfer_method": "local_file",
                    "upload_file_id": file_id,
                    "filename": file_name
                }
            ]
        }

        text = self._post_and_collect_stream_answer(data, "文件分析对话", 120)
        if text:
            print(f"\n=== AI 回答 ===\n{_pretty_json_text_for_display(text)}\n================\n")
            return text
        return None

    def analyze_audio(
        self,
        audio_path: str,
        query: str = "请分析这段音频的质量和内容",
        *,
        audio_eval_prompt: Optional[str] = None,
        selected_model: Optional[str] = None,
    ) -> Optional[str]:
        """
        分析音频文件
        
        Args:
            audio_path: 音频文件路径
            query: 分析问题
            selected_model: Dify ``inputs.selected_model``（开始表单必填类应用）
            
        Returns:
            str: AI 的分析结果
        """
        path = Path(audio_path)
        if not path.exists():
            raise FileNotFoundError(f"音频文件不存在：{audio_path}")
        
        # 检查文件格式
        supported_formats = {".mp3", ".wav", ".m4a", ".flac", ".ogg"}
        if path.suffix.lower() not in supported_formats:
            print(f"警告: {path.suffix} 可能不被支持，支持的格式: {supported_formats}")
        
        audio_mime_map = {
            ".mp3": "audio/mpeg",
            ".wav": "audio/wav",
            ".m4a": "audio/mp4",
            ".aac": "audio/aac",
            ".flac": "audio/flac",
            ".ogg": "audio/ogg",
        }
        upload_mime = audio_mime_map.get(path.suffix.lower(), "audio/wav")
        upload_result = self._upload_audio_for_chat(path, upload_mime)

        file_id = upload_result.get("id")
        file_name = upload_result.get("name", path.name)

        if not file_id:
            raise RuntimeError("Dify 上传返回中缺少文件 id")

        _post = _float_env("DIFY_SINGLE_AUDIO_POST_UPLOAD_DELAY_SEC", _DEFAULT_SINGLE_POST_UPLOAD)
        if _post > 0:
            self._emit(f"[Dify] 音频已上传，等待 {_post:.1f}s 后再发起对话（避免附件未就绪）…")
            _live_scoring_detail(f"音频已上传，等待 {_post:.0f}s 后发起评分…")
            time.sleep(_post)
        else:
            _live_scoring_detail("音频已上传，正在发起评分对话…")

        # 使用上传的音频进行对话
        data = {
            "inputs": _dify_chat_form_inputs(
                query,
                audio_eval_prompt_explicit=audio_eval_prompt,
                selected_model_explicit=selected_model,
            ),
            "query": query,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": self.user,
            "files": [
                {
                    "type": "audio",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id,
                    "filename": file_name
                }
            ]
        }
        
        text = self._post_and_collect_stream_answer(
            data, log_label=f"音频分析「{path.name}」", timeout=300
        )
        if text:
            print(f"\n=== AI 音频分析结果 ===\n{_pretty_json_text_for_display(text)}\n======================\n")
            return text
        return None


    def upload_audios_for_stimulus_compare(
        self,
        audio_paths: Sequence[str],
        device_slot_labels: Sequence[str],
    ) -> list[dict]:
        """仅上传刺激比较所需附件，返回 ``chat-messages`` 的 ``files`` 列表；可与 ``chat_stimulus_compare_with_uploaded_files`` 组合，避免解析重试时的重复上传。"""
        paths = [Path(p) for p in audio_paths]
        if len(paths) != len(device_slot_labels):
            raise ValueError("audio_paths 与 device_slot_labels 数量不一致")
        if not paths:
            return []

        audio_mime_map = {
            ".mp3": "audio/mpeg",
            ".wav": "audio/wav",
            ".m4a": "audio/mp4",
            ".aac": "audio/aac",
            ".flac": "audio/flac",
            ".ogg": "audio/ogg",
        }

        _gap = _float_env("DIFY_PER_AUDIO_UPLOAD_GAP_SEC", _DEFAULT_PER_UPLOAD_GAP)
        file_entries: list[dict] = []
        _n_paths = len(paths)
        for _pi, (path, label) in enumerate(zip(paths, device_slot_labels), start=1):
            if not path.exists():
                raise FileNotFoundError(f"音频不存在：{path}")
            self._emit(
                f"[Dify] 刺激比较：本音源第 {_pi}/{_n_paths} 路附件（{label}）"
                f"「{path.name[:72]}…」开始规范化与上传"
            )
            upload_mime = audio_mime_map.get(path.suffix.lower(), "audio/wav")
            upload_result = self._upload_audio_for_chat(path, upload_mime)
            file_id = upload_result.get("id")
            file_name = upload_result.get("name", path.name)
            if not file_id:
                raise RuntimeError(f"Dify 上传返回缺少文件 id：{path.name}")
            file_entries.append(
                {
                    "type": "audio",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id,
                    "filename": file_name,
                }
            )
            if _gap > 0:
                time.sleep(_gap)

        _fids = [str(e.get("upload_file_id", ""))[:12] for e in file_entries]
        _agent_dbg_upload(
            "H7",
            "difyclient.upload_audios_for_stimulus_compare:done",
            "files_uploaded",
            {"n_files": len(file_entries), "file_id_prefixes": _fids},
        )
        return file_entries

    def chat_stimulus_compare_with_uploaded_files(
        self,
        file_entries: Sequence[dict],
        *,
        stimulus_label: str,
        device_slot_labels: Sequence[str],
        extra_instruction: str = "",
        dut_attachment_index: int = 0,
        ref_attachment_index: int = 1,
        comparison_variant: str = "same_session",
        audio_eval_prompt: Optional[str] = None,
        selected_model: Optional[str] = None,
        post_upload_chat_delay_sec: Optional[float] = None,
        log_reuse_hint: bool = False,
        prompt_mode: str = "builtin",
    ) -> Optional[str]:
        """对已上传的 ``upload_file_id`` 发起刺激比较对话（不再次上传）。"""
        entries = [dict(x) for x in file_entries]
        if len(entries) != len(device_slot_labels):
            raise ValueError("file_entries 与 device_slot_labels 数量不一致")
        if not entries:
            return None

        if post_upload_chat_delay_sec is not None:
            _settle = float(post_upload_chat_delay_sec)
        else:
            _settle = _float_env(
                "DIFY_STIMULUS_POST_UPLOAD_CHAT_DELAY_SEC",
                _DEFAULT_STIMULUS_POST_UPLOAD,
            )

        _fids2 = [str(e.get("upload_file_id", ""))[:12] for e in entries]
        _agent_dbg_upload(
            "H7b",
            "difyclient.chat_stimulus_compare_with_uploaded_files:before_settle",
            "chat_with_existing_ids",
            {
                "n_files": len(entries),
                "file_id_prefixes": _fids2,
                "settle_sec": float(_settle),
                "reuse": bool(log_reuse_hint),
            },
        )

        if log_reuse_hint:
            self._emit(
                f"[Dify] 刺激比较：复用已上传的 {len(entries)} 路附件（不再上传），"
                f"等待 {_settle:.1f}s 后发起评分…"
            )
        else:
            self._emit(
                f"[Dify] 刺激比较：多路音频已上传，等待 {_settle:.1f}s 后再发起评分（附件就绪）…"
            )
        if _settle > 0:
            time.sleep(_settle)
        _live_scoring_detail(
            f"刺激比较「{(stimulus_label or '')[:32]}」：附件就绪，正在发起评分…"
        )

        _pm = (prompt_mode or "builtin").strip().lower()
        query = _build_stimulus_compare_query(
            extra_instruction=extra_instruction,
            stimulus_label=stimulus_label,
            device_slot_labels=device_slot_labels,
            comparison_variant=comparison_variant,
            dut_attachment_index=dut_attachment_index,
            ref_attachment_index=ref_attachment_index,
            prompt_mode=_pm,
        )
        if _pm == "final":
            self._emit(
                f"[Dify] 刺激比较：使用「仅最终提示词」模式（query 约 {len(query)} 字，"
                "未拼接内置长模板；未传 inputs.audio_eval_prompt）"
            )

        data = {
            "inputs": _dify_chat_form_inputs(
                query,
                audio_eval_prompt_explicit=audio_eval_prompt,
                selected_model_explicit=selected_model,
                omit_audio_eval_prompt=_pm == "final",
            ),
            "query": query,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": self.user,
            "files": entries,
        }

        _st = (stimulus_label or "").strip()
        _cmp_label = (
            f"刺激比较「{_st[:40]}…」" if len(_st) > 40 else f"刺激比较「{_st or '未命名音源'}」"
        )
        _agent_dbg_upload(
            "H8",
            "difyclient.chat_stimulus_compare_with_uploaded_files:before_chat",
            "post_chat_messages",
            {
                "query_len": len(query),
                "n_files": len(entries),
                "cmp_label": _cmp_label[:80],
            },
        )
        text = self._post_and_collect_stream_answer(data, _cmp_label, 600)
        if text:
            pretty = _pretty_json_text_for_display(text)
            self._emit(
                f"[Dify] 刺激比较：本轮模型正文 {len(text)} 字符（重试时会出现多行；"
                "全文仅排查时输出：环境变量 SPEAKER_DIFY_PRINT_STIMULUS_JSON=0）"
            )
            if (os.environ.get("SPEAKER_DIFY_PRINT_STIMULUS_JSON", "") or "").strip().lower() in (
                "1",
                "true",
                "yes",
                "on",
            ):
                print(f"\n=== AI 刺激比较结果 ===\n{pretty}\n======================\n")
            return text
        return None

    def analyze_audios_stimulus_compare(
        self,
        audio_paths: Sequence[str],
        device_slot_labels: Sequence[str],
        stimulus_label: str,
        extra_instruction: str = "",
        dut_attachment_index: int = 0,
        ref_attachment_index: int = 1,
        comparison_variant: str = "same_session",
        *,
        audio_eval_prompt: Optional[str] = None,
        selected_model: Optional[str] = None,
        post_upload_chat_delay_sec: Optional[float] = None,
    ) -> Optional[str]:
        """
        刺激比较法：同一音源下多路麦克风录音，作为多个 audio 附件一次性送评。
        默认第 dut_attachment_index+1 路为【被测】、第 ref_attachment_index+1 路为【对比】，
        五维分差为被测相对对比（仅允许整数 -3～+3）。

        comparison_variant:
        - ``same_session``：同一次测试流程内顺序录多机（默认）。
        - ``cross_session``：两次会话各录一台，评分时对齐音源；提示词中会强调
          **与同会话完全相同的评分规则与 JSON**，不得改用其它量表。

        audio_eval_prompt:
        - 传入 Dify ``inputs.audio_eval_prompt``（部分应用在开始表单中必填）；默认见 ``_dify_inputs_audio_eval_prompt``。
        selected_model:
        - 传入 Dify ``inputs.selected_model``；默认见 ``_dify_inputs_selected_model``。

        post_upload_chat_delay_sec:
        - 若指定，则覆盖环境变量 ``DIFY_STIMULUS_POST_UPLOAD_CHAT_DELAY_SEC``（用于拒答重试时加长等待）。

        实现上为 ``upload_audios_for_stimulus_compare`` + ``chat_stimulus_compare_with_uploaded_files``；
        解析重试栈（见 ``scoring``）会对同一批 ``upload_file_id`` 仅再次发起对话、不重复上传。
        """
        # #region agent log
        _agent_dbg_upload(
            "H6",
            "difyclient.analyze_audios_stimulus_compare:entry",
            "stimulus_compare_start",
            {
                "n_paths": len(list(audio_paths)),
                "stimulus_snip": (stimulus_label or "")[:120],
                "post_delay_override": post_upload_chat_delay_sec,
            },
        )
        # #endregion
        fe = self.upload_audios_for_stimulus_compare(audio_paths, device_slot_labels)
        return self.chat_stimulus_compare_with_uploaded_files(
            fe,
            stimulus_label=stimulus_label,
            device_slot_labels=device_slot_labels,
            extra_instruction=extra_instruction,
            dut_attachment_index=dut_attachment_index,
            ref_attachment_index=ref_attachment_index,
            comparison_variant=comparison_variant,
            audio_eval_prompt=audio_eval_prompt,
            selected_model=selected_model,
            post_upload_chat_delay_sec=post_upload_chat_delay_sec,
            log_reuse_hint=False,
        )

    def batch_analyze_audio(self, audio_dir: str, query: str = "分析这段音频的质量", output_file: str = "analysis_results.json") -> dict:
        """
        批量分析音频文件
        
        Args:
            audio_dir: 音频文件目录
            query: 分析问题
            output_file: 结果保存文件
            
        Returns:
            dict: 分析结果字典 {文件名: 分析结果}
        """
        from pathlib import Path as PathLib
        
        audio_path = PathLib(audio_dir)
        if not audio_path.exists():
            raise FileNotFoundError(f"目录不存在：{audio_dir}")
        
        # 获取所有音频文件
        audio_files = sorted([
            f for f in audio_path.iterdir() 
            if f.suffix.lower() in {".mp3", ".wav", ".m4a", ".flac", ".ogg"}
        ])
        
        if not audio_files:
            print(f"在 {audio_dir} 中未找到音频文件")
            return {}
        
        print(f"找到 {len(audio_files)} 个音频文件，开始批量分析...\n")
        
        results = {}
        
        for i, audio_file in enumerate(audio_files, 1):
            print(f"\n{'='*60}")
            print(f"[{i}/{len(audio_files)}] 分析: {audio_file.name}")
            print(f"{'='*60}")
            
            try:
                result = self.analyze_audio(str(audio_file), query)
                results[audio_file.name] = result
                
                # 每次分析后稍作停顿，避免 API 限流
                time.sleep(2)
                
            except Exception as e:
                print(f"处理 {audio_file.name} 时出错: {e}")
                results[audio_file.name] = f"ERROR: {str(e)}"
        
        # 保存结果到 JSON 文件
        if output_file:
            output_path = PathLib(output_file)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            print(f"\n分析结果已保存到: {output_file}")
        
        return results

    def upload_file_generic(
        self,
        file_path: str,
        mime_type: str = "text/plain",
        *,
        upload_label: Optional[str] = None,
    ) -> dict:
        """
        通用文件上传方法（含自动重试）。

        Args:
            file_path: 本地文件路径
            mime_type: 文件的 MIME 类型
            upload_label: 日志中展示的标签（默认用文件名）

        Returns:
            dict: 包含文件信息的字典，如 {"id": "xxx", "name": "filename.jpg"}
        """
        path = Path(file_path)
        label = upload_label or f"文件 {path.name}"
        return self._upload_file_multipart_with_retry(path, mime_type, label)


# #region agent log
_agent_dbg_upload(
    "BOOT",
    "difyclient:module",
    "module_loaded",
    {"difyclient_py": str(Path(__file__).resolve())[:500], "cwd": str(Path.cwd())},
)
# #endregion


if __name__ == '__main__':
    a = DifyClient()

    # 方式 1: 使用新方法 (推荐)
    # a.chat_with_image(
    #     r"C:\Users\Lenovo\Downloads\0_flesh_1 (5).jpg",
    #     "分析图片是否有异常问题，画面是清晰，界面字体大小正常，画面是否重叠，显示不全等异常，返回一个json格式的答案，格式为：{'是否异常': 'True/FAIL',异常现象:'xxxxx'}"
    #
    # )
    # a.analyze_file(r"C:\Users\Lenovo\Desktop\希倍思__XPK31_EVT_软件性能类_CPU专项测试报告_Fail_V1.0 (20241204).xlsx", "分析这份测试报告，里面是否存在测试fail项，需要结合测试一起审核，如果测试结果pass，但是测试数据不符合也需要提醒")

    # 方式 2: 纯文本对话
    # answer = a.rum("AI驱动的APP ui自动化测试实现方案")
    # print(answer)

    # 方式 3: 使用旧方法 (会自动调用新方法)
    # a.image_rum(r"C:\Users\Lenovo\Downloads\e6c67ab87761ea6467e0fcca7a6b45d2.jpeg")
